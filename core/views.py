import csv
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db.models import Q, Sum,Count
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse

from requests_app.forms import ServiceRequestForm, ClientPaymentProofForm, CancelRequestForm
from requests_app.models import ServiceRequest, RequestActivity

from reviews.forms import ReviewForm
from reviews.models import Review
from services.models import ServiceCategory

from technicians.models import TechnicianProfile
from technicians.forms import TechnicianProfileForm
from django.db import models


"""Admins"""
from django.contrib.admin.views.decorators import staff_member_required
from accounts.models import User
from requests_app.admin_forms import (
    AssignTechnicianForm,
    UpdateRequestStatusForm,
    UpdateAdminNotesForm,
    UpdateFinalPriceForm,
    UpdatePaymentForm,
    UpdateTechnicianPayoutForm,

)




def log_request_activity(service_request, action, message, user=None):
    RequestActivity.objects.create(
        service_request=service_request,
        action=action,
        message=message,
        performed_by=user if user and user.is_authenticated else None,
    )

def home(request):
    if request.method == "POST":
        form = ServiceRequestForm(request.POST, request.FILES)

        if form.is_valid():
            service_request = form.save(commit=False)

            if request.user.is_authenticated:
                service_request.client = request.user

            if service_request.category:
                service_request.estimated_price = service_request.category.starting_price

            service_request.save()

            messages.success(request, "Your request has been submitted successfully.")
            return redirect("request_success")
        else:
            messages.error(request, "Please correct the errors in the form.")

    else:
        form = ServiceRequestForm()

    service_categories = ServiceCategory.objects.filter(
        is_active=True
    ).order_by("display_order", "name")[:8]

    return render(request, "home.html", {
        "form": form,
        "service_categories": service_categories,
    })


def request_success(request):
    return render(request, "request_success.html")


@login_required
def client_dashboard(request):
    requests = ServiceRequest.objects.filter(
        client=request.user
    ).select_related(
        "category",
        "assigned_technician",
        "assigned_technician__user",
    )

    return render(request, "client/dashboard.html", {
        "requests": requests
    })


@login_required
def technician_dashboard(request):
    technician_profile = getattr(request.user, "technician_profile", None)

    assigned_requests = ServiceRequest.objects.none()
    total_technician_earnings = ServiceRequest.objects.filter(
    assigned_technician=technician_profile,
        status=ServiceRequest.STATUS_COMPLETED,
    ).aggregate(
        total=Sum("technician_earning")
    )["total"] or 0
    pending_requests = ServiceRequest.objects.filter(
        status=ServiceRequest.STATUS_PENDING
    ).select_related("category", "client")

    if technician_profile:
        assigned_requests = ServiceRequest.objects.filter(
            assigned_technician=technician_profile
        ).select_related("category", "client")

    paid_technician_earnings = ServiceRequest.objects.filter(
        assigned_technician=technician_profile,
        status=ServiceRequest.STATUS_COMPLETED,
        technician_payout_status=ServiceRequest.PAYOUT_PAID,
    ).aggregate(
        total=Sum("technician_earning")
    )["total"] or 0

    unpaid_technician_earnings = total_technician_earnings - paid_technician_earnings

    return render(request, "technician/dashboard.html", {
        "technician_profile": technician_profile,
        "assigned_requests": assigned_requests,
        "pending_requests": pending_requests,
        "total_technician_earnings": total_technician_earnings,
        "paid_technician_earnings": paid_technician_earnings,
        "unpaid_technician_earnings": unpaid_technician_earnings,
    })

@login_required
def accept_request(request, request_id):
    technician_profile = getattr(request.user, "technician_profile", None)

    if not technician_profile:
        messages.error(request, "You do not have a technician profile.")
        return redirect("technician_dashboard")

    if not technician_profile.is_verified:
        messages.error(request, "Your technician account is not verified yet. Please wait for admin approval.")
        return redirect("technician_dashboard")

    service_request = get_object_or_404(
        ServiceRequest,
        id=request_id,
        status=ServiceRequest.STATUS_PENDING
    )

    service_request.assigned_technician = technician_profile
    service_request.status = ServiceRequest.STATUS_ACCEPTED
    service_request.save()

    messages.success(request, "Job accepted successfully.")
    return redirect("technician_dashboard")

@login_required
def start_request(request, request_id):
    technician_profile = getattr(request.user, "technician_profile", None)

    if not technician_profile:
        messages.error(request, "You do not have a technician profile.")
        return redirect("technician_dashboard")

    service_request = get_object_or_404(
    ServiceRequest,
    id=request_id,
    assigned_technician=technician_profile,
    status__in=[
        ServiceRequest.STATUS_ASSIGNED,
        ServiceRequest.STATUS_ACCEPTED,
    ],
    )
    service_request.status = ServiceRequest.STATUS_IN_PROGRESS
    service_request.save()

    messages.success(request, "Job marked as in progress.")
    return redirect("technician_dashboard")


@login_required
def complete_request(request, request_id):
    technician_profile = getattr(request.user, "technician_profile", None)

    if not technician_profile:
        messages.error(request, "You do not have a technician profile.")
        return redirect("technician_dashboard")

    service_request = get_object_or_404(
        ServiceRequest,
        id=request_id,
        assigned_technician=technician_profile,
        status=ServiceRequest.STATUS_IN_PROGRESS,
    )

    service_request.status = ServiceRequest.STATUS_COMPLETED
    service_request.completed_at = timezone.now()
    service_request.save()

    messages.success(request, "Job marked as completed.")
    return redirect("technician_dashboard")

@login_required
def leave_review(request, request_id):
    service_request = get_object_or_404(
        ServiceRequest,
        id=request_id,
        client=request.user,
        status=ServiceRequest.STATUS_COMPLETED,
    )

    if not service_request.assigned_technician:
        messages.error(request, "This request has no assigned technician.")
        return redirect("client_dashboard")

    existing_review = Review.objects.filter(
        request=service_request
    ).first()

    if existing_review:
        messages.info(request, "You have already reviewed this job.")
        return redirect("client_dashboard")

    if request.method == "POST":
        form = ReviewForm(request.POST)

        if form.is_valid():
            review = form.save(commit=False)
            review.client = request.user
            review.technician = service_request.assigned_technician
            review.request = service_request
            review.save()

            messages.success(request, "Review submitted successfully.")
            return redirect("client_dashboard")
    else:
        form = ReviewForm()

    return render(request, "client/leave_review.html", {
        "form": form,
        "service_request": service_request,
    })   

@login_required
def client_request_detail(request, request_id):
    service_request = get_object_or_404(
        ServiceRequest.objects.select_related(
            "category",
            "assigned_technician",
            "assigned_technician__user",
            "review",
        ),
        id=request_id,
        client=request.user,
    )

    can_client_cancel = service_request.status in [
        ServiceRequest.STATUS_PENDING,
        ServiceRequest.STATUS_ASSIGNED,
        ServiceRequest.STATUS_ACCEPTED,
    ]

    return render(request, "client/request_detail.html", {
        "service_request": service_request,
        "can_client_cancel": can_client_cancel,

    })    

@login_required
def edit_technician_profile(request):
    technician_profile = getattr(request.user, "technician_profile", None)

    if not technician_profile:
        messages.error(request, "You do not have a technician profile.")
        return redirect("technician_dashboard")

    if request.method == "POST":
        form = TechnicianProfileForm(
            request.POST,
            request.FILES,
            instance=technician_profile
        )
        if form.is_valid():
            form.save()
            messages.success(request, "Technician profile updated successfully.")
            return redirect("technician_dashboard")
    else:
        form = TechnicianProfileForm(instance=technician_profile)

    return render(request, "technician/edit_profile.html", {
        "form": form,
        "technician_profile": technician_profile,
    })

@login_required
def technician_request_detail(request, request_id):
    technician_profile = getattr(request.user, "technician_profile", None)

    if not technician_profile:
        messages.error(request, "You do not have a technician profile.")
        return redirect("technician_dashboard")

    service_request = get_object_or_404(
        ServiceRequest.objects.select_related(
            "category",
            "client",
            "assigned_technician",
            "assigned_technician__user",
        ),
        id=request_id,
    )

    # Technician can view:
    # 1. Pending jobs not yet assigned
    # 2. Jobs assigned to them
    if service_request.status != ServiceRequest.STATUS_PENDING and service_request.assigned_technician != technician_profile:
        messages.error(request, "You are not allowed to view this job.")
        return redirect("technician_dashboard")

    return render(request, "technician/request_detail.html", {
    "service_request": service_request,
    "technician_profile": technician_profile,
})

@staff_member_required
def admin_dashboard(request):
    total_requests = ServiceRequest.objects.count()
    pending_requests_count = ServiceRequest.objects.filter(
        status=ServiceRequest.STATUS_PENDING
    ).count()
    accepted_requests_count = ServiceRequest.objects.filter(
        status=ServiceRequest.STATUS_ACCEPTED
    ).count()
    in_progress_requests_count = ServiceRequest.objects.filter(
        status=ServiceRequest.STATUS_IN_PROGRESS
    ).count()
    completed_requests_count = ServiceRequest.objects.filter(
        status=ServiceRequest.STATUS_COMPLETED
    ).count()

    total_clients = User.objects.filter(role=User.ROLE_CLIENT).count()
    total_technicians = TechnicianProfile.objects.count()
    verified_technicians = TechnicianProfile.objects.filter(is_verified=True).count()
    unverified_technicians = TechnicianProfile.objects.filter(is_verified=False).count()

    pending_payment_proofs = ServiceRequest.objects.filter(
        payment_proof_status=ServiceRequest.PROOF_PENDING
    ).count()

    approved_payment_proofs = ServiceRequest.objects.filter(
        payment_proof_status=ServiceRequest.PROOF_APPROVED
    ).count()

    rejected_payment_proofs = ServiceRequest.objects.filter(
        payment_proof_status=ServiceRequest.PROOF_REJECTED
    ).count()

    not_submitted_payment_proofs = ServiceRequest.objects.filter(
        payment_proof_status=ServiceRequest.PROOF_NOT_SUBMITTED
    ).count()

    total_final_value = ServiceRequest.objects.aggregate(
        total=Sum("final_price")
    )["total"] or 0

    total_amount_paid = ServiceRequest.objects.aggregate(
        total=Sum("amount_paid")
    )["total"] or 0

    outstanding_balance = total_final_value - total_amount_paid

    paid_jobs_count = ServiceRequest.objects.filter(
        payment_status=ServiceRequest.PAYMENT_PAID
    ).count()

    latest_requests = ServiceRequest.objects.select_related(
        "category",
        "client",
        "assigned_technician",
        "assigned_technician__user",
    ).order_by("-created_at")[:10]

    latest_technicians = TechnicianProfile.objects.select_related(
        "user"
    ).order_by("-created_at")[:8]

    pending_proof_requests = ServiceRequest.objects.select_related(
    "category",
    "client",
    "assigned_technician",
    "assigned_technician__user",
    ).filter(
        payment_proof_status=ServiceRequest.PROOF_PENDING
    ).order_by("-updated_at")[:8]

    unpaid_payout_requests = ServiceRequest.objects.select_related(
        "category",
        "client",
        "assigned_technician",
        "assigned_technician__user",
    ).filter(
        status=ServiceRequest.STATUS_COMPLETED,
        technician_payout_status=ServiceRequest.PAYOUT_UNPAID,
        technician_earning__gt=0,
    ).order_by("-completed_at")[:10]

    total_technician_earnings = ServiceRequest.objects.aggregate(
        total=Sum("technician_earning")
    )["total"] or 0

    total_technician_paid_out = ServiceRequest.objects.filter(
        technician_payout_status=ServiceRequest.PAYOUT_PAID
    ).aggregate(
        total=Sum("technician_earning")
    )["total"] or 0

    total_technician_unpaid = total_technician_earnings - total_technician_paid_out

    admin_alerts = []

    if pending_requests_count > 0:
        admin_alerts.append({
            "title": "Pending service requests",
            "message": f"{pending_requests_count} request(s) are waiting for technician assignment.",
            "type": "warning",
            "url": "/dashboard/admin/requests/?status=pending",
            "action": "Assign now",
        })

    if pending_payment_proofs > 0:
        admin_alerts.append({
            "title": "Payment proofs need review",
            "message": f"{pending_payment_proofs} payment proof(s) are waiting for approval.",
            "type": "warning",
            "url": "/dashboard/admin/",
            "action": "Review proofs",
        })

    if total_technician_unpaid > 0:
        admin_alerts.append({
            "title": "Technician payouts pending",
            "message": f"UGX {total_technician_unpaid:,.0f} is still unpaid to technicians.",
            "type": "danger",
            "url": "/dashboard/admin/",
            "action": "Review payouts",
        })

    if unverified_technicians > 0:
        admin_alerts.append({
            "title": "Technicians waiting verification",
            "message": f"{unverified_technicians} technician profile(s) need admin verification.",
            "type": "info",
            "url": "/admin/technicians/technicianprofile/",
            "action": "Verify",
        })

    if outstanding_balance > 0:
        admin_alerts.append({
            "title": "Outstanding client balances",
            "message": f"UGX {outstanding_balance:,.0f} is still outstanding from clients.",
            "type": "danger",
            "url": "/dashboard/admin/requests/",
            "action": "View balances",
        })

    return render(request, "admin_dashboard/dashboard.html", {
        "total_requests": total_requests,
        "pending_requests_count": pending_requests_count,
        "accepted_requests_count": accepted_requests_count,
        "in_progress_requests_count": in_progress_requests_count,
        "completed_requests_count": completed_requests_count,
        "total_clients": total_clients,
        "total_technicians": total_technicians,
        "verified_technicians": verified_technicians,
        "unverified_technicians": unverified_technicians,
        "latest_requests": latest_requests,
        "latest_technicians": latest_technicians,
        "pending_payment_proofs": pending_payment_proofs,
        "approved_payment_proofs": approved_payment_proofs,
        "rejected_payment_proofs": rejected_payment_proofs,
        "not_submitted_payment_proofs": not_submitted_payment_proofs,
        "total_final_value": total_final_value,
        "total_amount_paid": total_amount_paid,
        "outstanding_balance": outstanding_balance,
        "paid_jobs_count": paid_jobs_count,
        "pending_proof_requests": pending_proof_requests,
        "unpaid_payout_requests": unpaid_payout_requests,
        "total_technician_earnings": total_technician_earnings,
        "total_technician_paid_out": total_technician_paid_out,
        "total_technician_unpaid": total_technician_unpaid,
        "admin_alerts": admin_alerts,

    })

@staff_member_required
def admin_assign_request(request, request_id):
    service_request = get_object_or_404(
        ServiceRequest.objects.select_related(
            "category",
            "client",
            "assigned_technician",
            "assigned_technician__user",
        ),
        id=request_id,
    )

    if request.method == "POST":
        form = AssignTechnicianForm(request.POST)

        if form.is_valid():
            technician = form.cleaned_data["technician"]

            service_request.assigned_technician = technician
            service_request.status = ServiceRequest.STATUS_ASSIGNED
            service_request.save()

            log_request_activity(
                service_request,
                RequestActivity.ACTION_ASSIGNED,
                f"Admin assigned technician {technician.user.username}.",
                request.user,
            )


            messages.success(request, "Technician assigned successfully.")
            return redirect("admin_dashboard")
    else:
        form = AssignTechnicianForm()

    return render(request, "admin_dashboard/assign_request.html", {
        "form": form,
        "service_request": service_request,
    })


@staff_member_required
def admin_request_detail(request, request_id):
    service_request = get_object_or_404(
        ServiceRequest.objects.select_related(
            "category",
            "client",
            "assigned_technician",
            "assigned_technician__user",
            "review",
        ),
        id=request_id,
    )

    status_form = UpdateRequestStatusForm(initial={
        "status": service_request.status
    })

    notes_form = UpdateAdminNotesForm(initial={
        "admin_notes": service_request.admin_notes
    })

    price_form = UpdateFinalPriceForm(initial={
        "final_price": service_request.final_price,
        "price_note": service_request.price_note,
        "platform_commission_percent": service_request.platform_commission_percent,

    })

    payout_form = UpdateTechnicianPayoutForm(initial={
        "technician_payout_status": service_request.technician_payout_status,
        "technician_payout_method": service_request.technician_payout_method,
        "technician_payout_note": service_request.technician_payout_note,
    })

    payment_form = UpdatePaymentForm(initial={
        "payment_status": service_request.payment_status,
        "payment_method": service_request.payment_method,
        "amount_paid": service_request.amount_paid,
        "payment_note": service_request.payment_note,
        "payment_proof_status": service_request.payment_proof_status,
    })

    activities = service_request.activities.select_related(
        "performed_by"
    ).order_by("-created_at")

    return render(request, "admin_dashboard/request_detail.html", {
        "service_request": service_request,
        "status_form": status_form,
        "notes_form": notes_form,
        "price_form": price_form,
        "payment_form": payment_form,
        "payout_form": payout_form,
        "activities": activities,
        "activities_count": activities.count(),
    })


@staff_member_required
def admin_update_request_status(request, request_id):
    service_request = get_object_or_404(ServiceRequest, id=request_id)

    if request.method == "POST":
        form = UpdateRequestStatusForm(request.POST)

        if form.is_valid():
            old_status = service_request.status
            new_status = form.cleaned_data["status"]

            service_request.status = new_status

            # If admin marks completed, set completed date
            if new_status == ServiceRequest.STATUS_COMPLETED and not service_request.completed_at:
                service_request.completed_at = timezone.now()

            # If admin moves away from completed, clear completed date
            if new_status != ServiceRequest.STATUS_COMPLETED:
                service_request.completed_at = None

            # If admin reopens a cancelled request, clear cancellation reason
            if old_status == ServiceRequest.STATUS_CANCELLED and new_status != ServiceRequest.STATUS_CANCELLED:
                service_request.cancellation_reason = ""

            service_request.save()

            log_request_activity(
                service_request,
                RequestActivity.ACTION_STATUS_UPDATED,
                f"Admin changed status from {old_status} to {new_status}.",
                request.user,
            )

            messages.success(request, "Request status updated successfully.")
            return redirect("admin_request_detail", request_id=service_request.id)

        messages.error(request, form.errors)

    messages.error(request, "Invalid status update.")
    return redirect("admin_request_detail", request_id=service_request.id)




@staff_member_required
def admin_update_final_price(request, request_id):
    service_request = get_object_or_404(ServiceRequest, id=request_id)

    if request.method == "POST":
        form = UpdateFinalPriceForm(request.POST)

        if form.is_valid():
            service_request.final_price = form.cleaned_data["final_price"]
            service_request.price_note = form.cleaned_data["price_note"]
            service_request.platform_commission_percent = form.cleaned_data["platform_commission_percent"]
            service_request.calculate_earnings()
            service_request.save()

            log_request_activity(
                service_request,
                RequestActivity.ACTION_PRICE_UPDATED,
                f"Admin updated final price to UGX {service_request.final_price}.",
                request.user,
            )

            messages.success(request, "Final price updated successfully.")
            return redirect("admin_request_detail", request_id=service_request.id)

    messages.error(request, "Invalid final price update.")
    return redirect("admin_request_detail", request_id=service_request.id)

@staff_member_required
def admin_update_request_notes(request, request_id):
    service_request = get_object_or_404(ServiceRequest, id=request_id)

    if request.method == "POST":
        form = UpdateAdminNotesForm(request.POST)

        if form.is_valid():
            service_request.admin_notes = form.cleaned_data["admin_notes"]
            service_request.save()

            log_request_activity(
                service_request,
                RequestActivity.ACTION_NOTES_UPDATED,
                "Admin internal notes were updated.",
                request.user,
            )

            messages.success(request, "Admin notes updated successfully.")
            return redirect("admin_request_detail", request_id=service_request.id)

    messages.error(request, "Invalid notes update.")
    return redirect("admin_request_detail", request_id=service_request.id)   

@staff_member_required
def admin_update_payment(request, request_id):
    service_request = get_object_or_404(ServiceRequest, id=request_id)

    if request.method == "POST":
        form = UpdatePaymentForm(request.POST, request.FILES)

        if form.is_valid():
            service_request.payment_status = form.cleaned_data["payment_status"]
            service_request.payment_method = form.cleaned_data["payment_method"]
            service_request.amount_paid = form.cleaned_data["amount_paid"]
            service_request.payment_note = form.cleaned_data["payment_note"]
            service_request.payment_proof_status = form.cleaned_data["payment_proof_status"]
            if form.cleaned_data.get("payment_proof"):
                service_request.payment_proof = form.cleaned_data["payment_proof"]
            service_request.save()

            log_request_activity(
                service_request,
                RequestActivity.ACTION_PAYMENT_UPDATED,
                "Admin updated payment details.",
                request.user,
            )

            messages.success(request, "Payment details updated successfully.")
            return redirect("admin_request_detail", request_id=service_request.id)

    messages.error(request, "Invalid payment update.")
    return redirect("admin_request_detail", request_id=service_request.id)

@login_required
def client_upload_payment_proof(request, request_id):
    service_request = get_object_or_404(
        ServiceRequest,
        id=request_id,
        client=request.user,
    )

    if request.method == "POST":
        form = ClientPaymentProofForm(request.POST, request.FILES)

        if form.is_valid():
            service_request.payment_proof = form.cleaned_data["payment_proof"]
            service_request.payment_proof_status = ServiceRequest.PROOF_PENDING
            service_request.save()

            log_request_activity(
                service_request,
                RequestActivity.ACTION_PROOF_UPLOADED,
                "Client uploaded payment proof.",
                request.user,
            )

            messages.success(
                request,
                "Payment proof uploaded successfully. FixNet admin will review it."
            )
            return redirect("client_request_detail", request_id=service_request.id)
    else:
        form = ClientPaymentProofForm()

    return render(request, "client/upload_payment_proof.html", {
        "form": form,
        "service_request": service_request,
    })


@login_required
def client_cancel_request(request, request_id):
    service_request = get_object_or_404(
        ServiceRequest,
        id=request_id,
        client=request.user,
    )

    if service_request.status in [
        ServiceRequest.STATUS_IN_PROGRESS,
        ServiceRequest.STATUS_COMPLETED,
        ServiceRequest.STATUS_CANCELLED,
    ]:
        messages.error(request, "This request cannot be cancelled at this stage.")
        return redirect("client_request_detail", request_id=service_request.id)

    if request.method == "POST":
        form = CancelRequestForm(request.POST)

        if form.is_valid():
            service_request.cancellation_reason = form.cleaned_data["cancellation_reason"]
            service_request.status = ServiceRequest.STATUS_CANCELLED
            service_request.save()

            messages.success(request, "Request cancelled successfully.")
            return redirect("client_dashboard")
    else:
        form = CancelRequestForm()

    return render(request, "client/cancel_request.html", {
        "form": form,
        "service_request": service_request,
    })


@staff_member_required
def admin_cancel_request(request, request_id):
    service_request = get_object_or_404(ServiceRequest, id=request_id)

    if service_request.status == ServiceRequest.STATUS_COMPLETED:
        messages.error(request, "Completed requests should not be cancelled.")
        return redirect("admin_request_detail", request_id=service_request.id)

    service_request.status = ServiceRequest.STATUS_CANCELLED
    service_request.save()

    messages.success(request, "Request cancelled successfully.")
    return redirect("admin_request_detail", request_id=service_request.id)


@login_required
def technician_profile_detail(request, technician_id):
    technician_profile = get_object_or_404(
        TechnicianProfile.objects.select_related("user"),
        id=technician_id,
    )

    completed_jobs = ServiceRequest.objects.filter(
        assigned_technician=technician_profile,
        status=ServiceRequest.STATUS_COMPLETED,
    ).select_related(
        "category",
        "client",
    ).order_by("-completed_at")[:10]

    return render(request, "technician/profile_detail.html", {
        "technician_profile": technician_profile,
        "completed_jobs": completed_jobs,
    })

@staff_member_required
def admin_update_technician_payout(request, request_id):
    service_request = get_object_or_404(ServiceRequest, id=request_id)

    if request.method == "POST":
        form = UpdateTechnicianPayoutForm(request.POST)

        if form.is_valid():
            service_request.technician_payout_status = form.cleaned_data["technician_payout_status"]
            service_request.technician_payout_method = form.cleaned_data["technician_payout_method"]
            service_request.technician_payout_note = form.cleaned_data["technician_payout_note"]

            if service_request.technician_payout_status == ServiceRequest.PAYOUT_PAID:
                if not service_request.technician_payout_date:
                    service_request.technician_payout_date = timezone.now()
            else:
                service_request.technician_payout_date = None

            service_request.save()

            log_request_activity(
                service_request,
                RequestActivity.ACTION_PAYMENT_UPDATED,
                "Admin updated technician payout details.",
                request.user,
            )

            messages.success(request, "Technician payout updated successfully.")
            return redirect("admin_request_detail", request_id=service_request.id)

        messages.error(request, form.errors)

    messages.error(request, "Invalid technician payout update.")
    return redirect("admin_request_detail", request_id=service_request.id)

@login_required
def technician_job_history(request):
    technician_profile = get_object_or_404(
        TechnicianProfile,
        user=request.user,
    )

    jobs = ServiceRequest.objects.filter(
        assigned_technician=technician_profile
    ).select_related(
        "category",
        "client",
    ).order_by("-created_at")

    completed_jobs = jobs.filter(status=ServiceRequest.STATUS_COMPLETED)
    active_jobs = jobs.exclude(
        status__in=[
            ServiceRequest.STATUS_COMPLETED,
            ServiceRequest.STATUS_CANCELLED,
        ]
    )

    total_earnings = completed_jobs.aggregate(
        total=Sum("technician_earning")
    )["total"] or 0

    paid_earnings = completed_jobs.filter(
        technician_payout_status=ServiceRequest.PAYOUT_PAID
    ).aggregate(
        total=Sum("technician_earning")
    )["total"] or 0

    unpaid_earnings = total_earnings - paid_earnings

    return render(request, "technician/job_history.html", {
        "technician_profile": technician_profile,
        "jobs": jobs,
        "completed_jobs_count": completed_jobs.count(),
        "active_jobs_count": active_jobs.count(),
        "total_earnings": total_earnings,
        "paid_earnings": paid_earnings,
        "unpaid_earnings": unpaid_earnings,
    })

@login_required
def client_request_history(request):
    requests_qs = ServiceRequest.objects.filter(
        client=request.user
    ).select_related(
        "category",
        "assigned_technician",
        "assigned_technician__user",
        "review",
    ).order_by("-created_at")

    active_requests = requests_qs.exclude(
        status__in=[
            ServiceRequest.STATUS_COMPLETED,
            ServiceRequest.STATUS_CANCELLED,
        ]
    )

    completed_requests = requests_qs.filter(
        status=ServiceRequest.STATUS_COMPLETED
    )

    cancelled_requests = requests_qs.filter(
        status=ServiceRequest.STATUS_CANCELLED
    )

    total_final_value = requests_qs.aggregate(
        total=Sum("final_price")
    )["total"] or 0

    total_amount_paid = requests_qs.aggregate(
        total=Sum("amount_paid")
    )["total"] or 0

    total_balance = total_final_value - total_amount_paid

    return render(request, "client/request_history.html", {
        "requests_qs": requests_qs,
        "active_requests_count": active_requests.count(),
        "completed_requests_count": completed_requests.count(),
        "cancelled_requests_count": cancelled_requests.count(),
        "total_final_value": total_final_value,
        "total_amount_paid": total_amount_paid,
        "total_balance": total_balance,
    })

@staff_member_required
def admin_request_list(request):
    search_query = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "").strip()
    payment_status_filter = request.GET.get("payment_status", "").strip()
    proof_status_filter = request.GET.get("proof_status", "").strip()
    technician_filter = request.GET.get("technician", "").strip()

    requests_qs = ServiceRequest.objects.select_related(
        "category",
        "client",
        "assigned_technician",
        "assigned_technician__user",
    ).order_by("-created_at")

    if search_query:
        requests_qs = requests_qs.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(client_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(location__icontains=search_query)
        )

    if status_filter:
        requests_qs = requests_qs.filter(status=status_filter)

    if payment_status_filter:
        requests_qs = requests_qs.filter(payment_status=payment_status_filter)

    if proof_status_filter:
        requests_qs = requests_qs.filter(payment_proof_status=proof_status_filter)

    if technician_filter:
        requests_qs = requests_qs.filter(assigned_technician_id=technician_filter)

    technicians = TechnicianProfile.objects.select_related("user").order_by("user__username")

    status_options = [
        {
            "value": value,
            "label": label,
            "selected": status_filter == value,
        }
        for value, label in ServiceRequest.STATUS_CHOICES
    ]

    payment_status_options = [
        {
            "value": value,
            "label": label,
            "selected": payment_status_filter == value,
        }
        for value, label in ServiceRequest.PAYMENT_STATUS_CHOICES
    ]

    proof_status_options = [
        {
            "value": value,
            "label": label,
            "selected": proof_status_filter == value,
        }
        for value, label in ServiceRequest.PAYMENT_PROOF_STATUS_CHOICES
    ]

    technician_options = [
        {
            "value": str(tech.id),
            "label": tech.user.username,
            "selected": technician_filter == str(tech.id),
        }
        for tech in technicians
    ]

    paginator = Paginator(requests_qs, 15)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "admin_dashboard/request_list.html", {
        "requests_qs": page_obj,
        "page_obj": page_obj,
        "total_matching_requests": requests_qs.count(), 

        "search_query": search_query,
        "status_filter": status_filter,
        "payment_status_filter": payment_status_filter,
        "proof_status_filter": proof_status_filter,
        "technician_filter": technician_filter,

        "status_options": status_options,
        "payment_status_options": payment_status_options,
        "proof_status_options": proof_status_options,
        "technician_options": technician_options,
    })

@staff_member_required
def admin_live_stats(request):
    total_requests = ServiceRequest.objects.count()

    pending_requests_count = ServiceRequest.objects.filter(
        status=ServiceRequest.STATUS_PENDING
    ).count()

    accepted_requests_count = ServiceRequest.objects.filter(
        status=ServiceRequest.STATUS_ACCEPTED
    ).count()

    in_progress_requests_count = ServiceRequest.objects.filter(
        status=ServiceRequest.STATUS_IN_PROGRESS
    ).count()

    completed_requests_count = ServiceRequest.objects.filter(
        status=ServiceRequest.STATUS_COMPLETED
    ).count()

    pending_payment_proofs = ServiceRequest.objects.filter(
        payment_proof_status=ServiceRequest.PROOF_PENDING
    ).count()

    approved_payment_proofs = ServiceRequest.objects.filter(
        payment_proof_status=ServiceRequest.PROOF_APPROVED
    ).count()

    rejected_payment_proofs = ServiceRequest.objects.filter(
        payment_proof_status=ServiceRequest.PROOF_REJECTED
    ).count()

    not_submitted_payment_proofs = ServiceRequest.objects.filter(
        payment_proof_status=ServiceRequest.PROOF_NOT_SUBMITTED
    ).count()

    total_final_value = ServiceRequest.objects.aggregate(
        total=Sum("final_price")
    )["total"] or 0

    total_amount_paid = ServiceRequest.objects.aggregate(
        total=Sum("amount_paid")
    )["total"] or 0

    outstanding_balance = total_final_value - total_amount_paid

    total_technician_earnings = ServiceRequest.objects.aggregate(
        total=Sum("technician_earning")
    )["total"] or 0

    total_technician_paid_out = ServiceRequest.objects.filter(
        technician_payout_status=ServiceRequest.PAYOUT_PAID
    ).aggregate(
        total=Sum("technician_earning")
    )["total"] or 0

    total_technician_unpaid = total_technician_earnings - total_technician_paid_out

    return JsonResponse({
        "total_requests": total_requests,
        "pending_requests_count": pending_requests_count,
        "accepted_requests_count": accepted_requests_count,
        "in_progress_requests_count": in_progress_requests_count,
        "completed_requests_count": completed_requests_count,

        "pending_payment_proofs": pending_payment_proofs,
        "approved_payment_proofs": approved_payment_proofs,
        "rejected_payment_proofs": rejected_payment_proofs,
        "not_submitted_payment_proofs": not_submitted_payment_proofs,

        "total_final_value": float(total_final_value),
        "total_amount_paid": float(total_amount_paid),
        "outstanding_balance": float(outstanding_balance),

        "total_technician_earnings": float(total_technician_earnings),
        "total_technician_paid_out": float(total_technician_paid_out),
        "total_technician_unpaid": float(total_technician_unpaid),

        "charts": {
            "job_status": {
                "labels": ["Pending", "Accepted", "In Progress", "Completed"],
                "data": [
                    pending_requests_count,
                    accepted_requests_count,
                    in_progress_requests_count,
                    completed_requests_count,
                ],
            },
            "payment_collection": {
                "labels": ["Collected", "Outstanding"],
                "data": [
                    float(total_amount_paid),
                    float(outstanding_balance),
                ],
            },
            "proof_status": {
                "labels": ["Pending", "Approved", "Rejected", "Not Submitted"],
                "data": [
                    pending_payment_proofs,
                    approved_payment_proofs,
                    rejected_payment_proofs,
                    not_submitted_payment_proofs,
                ],
            },
            "technician_payout": {
                "labels": ["Paid Out", "Unpaid"],
                "data": [
                    float(total_technician_paid_out),
                    float(total_technician_unpaid),
                ],
            },
        }
    })

@staff_member_required
def admin_live_latest_requests(request):
    latest_requests = ServiceRequest.objects.select_related(
        "category",
        "client",
        "assigned_technician",
        "assigned_technician__user",
    ).order_by("-created_at")[:10]

    data = []

    for item in latest_requests:
        if item.client:
            client_name = item.client.username
        elif item.client_name:
            client_name = item.client_name
        else:
            client_name = "Public Client"

        if item.assigned_technician:
            technician_name = item.assigned_technician.user.username
            technician_url = f"/technicians/{item.assigned_technician.id}/"
        else:
            technician_name = "Not assigned"
            technician_url = ""

        if item.category:
            category_name = item.category.name
        else:
            category_name = "No category"

        data.append({
            "id": item.id,
            "title": item.title,
            "category": category_name,
            "client_name": client_name,
            "location": item.location,
            "technician_name": technician_name,
            "technician_url": technician_url,
            "status": item.status,
            "status_display": item.get_status_display(),
            "payment_proof_status": item.payment_proof_status,
            "detail_url": f"/dashboard/admin/requests/{item.id}/",
            "assign_url": f"/dashboard/admin/requests/{item.id}/assign/",
        })

    return JsonResponse({
        "latest_requests": data
    })

@staff_member_required
def admin_live_pending_proofs(request):
    pending_proof_requests = ServiceRequest.objects.select_related(
        "category",
        "client",
        "assigned_technician",
        "assigned_technician__user",
    ).filter(
        payment_proof_status=ServiceRequest.PROOF_PENDING
    ).order_by("-updated_at")[:10]

    data = []

    for item in pending_proof_requests:
        if item.client:
            client_name = item.client.username
        elif item.client_name:
            client_name = item.client_name
        else:
            client_name = "Public Client"

        data.append({
            "id": item.id,
            "title": item.title,
            "location": item.location,
            "client_name": client_name,
            "final_price": float(item.final_price or 0),
            "amount_paid": float(item.amount_paid or 0),
            "has_payment_proof": bool(item.payment_proof),
            "detail_url": f"/dashboard/admin/requests/{item.id}/",
        })

    return JsonResponse({
        "pending_proof_requests": data
    })

@staff_member_required
def admin_live_unpaid_payouts(request):
    unpaid_payout_requests = ServiceRequest.objects.select_related(
        "category",
        "client",
        "assigned_technician",
        "assigned_technician__user",
    ).filter(
        status=ServiceRequest.STATUS_COMPLETED,
        technician_payout_status=ServiceRequest.PAYOUT_UNPAID,
        technician_earning__gt=0,
    ).order_by("-completed_at")[:10]

    data = []

    for item in unpaid_payout_requests:
        if item.assigned_technician:
            technician_name = item.assigned_technician.user.username
            technician_url = f"/technicians/{item.assigned_technician.id}/"
        else:
            technician_name = "Not assigned"
            technician_url = ""

        data.append({
            "id": item.id,
            "title": item.title,
            "location": item.location,
            "technician_name": technician_name,
            "technician_url": technician_url,
            "final_price": float(item.final_price or 0),
            "technician_earning": float(item.technician_earning or 0),
            "completed_at": item.completed_at.strftime("%b %d, %Y") if item.completed_at else "Not recorded",
            "detail_url": f"/dashboard/admin/requests/{item.id}/",
        })

    return JsonResponse({
        "unpaid_payout_requests": data
    })

@staff_member_required
def admin_reports(request):
    today = timezone.localdate()

    date_filter = request.GET.get("date_filter", "this_month")
    start_date_input = request.GET.get("start_date", "")
    end_date_input = request.GET.get("end_date", "")

    start_date = None
    end_date = None

    if date_filter == "today":
        start_date = today
        end_date = today

    elif date_filter == "this_week":
        start_date = today - timedelta(days=today.weekday())
        end_date = today

    elif date_filter == "this_month":
        start_date = today.replace(day=1)
        end_date = today

    elif date_filter == "this_year":
        start_date = today.replace(month=1, day=1)
        end_date = today

    elif date_filter == "custom":
        if start_date_input:
            start_date = datetime.strptime(start_date_input, "%Y-%m-%d").date()

        if end_date_input:
            end_date = datetime.strptime(end_date_input, "%Y-%m-%d").date()

    requests_qs = ServiceRequest.objects.select_related(
        "category",
        "client",
        "assigned_technician",
        "assigned_technician__user",
    )

    if start_date:
        requests_qs = requests_qs.filter(created_at__date__gte=start_date)

    if end_date:
        requests_qs = requests_qs.filter(created_at__date__lte=end_date)

    completed_requests = requests_qs.filter(
        status=ServiceRequest.STATUS_COMPLETED
    )

    cancelled_requests = requests_qs.filter(
        status=ServiceRequest.STATUS_CANCELLED
    )

    total_requests = requests_qs.count()
    completed_jobs_count = completed_requests.count()
    cancelled_jobs_count = cancelled_requests.count()

    total_final_value = requests_qs.aggregate(
        total=Sum("final_price")
    )["total"] or 0

    total_amount_paid = requests_qs.aggregate(
        total=Sum("amount_paid")
    )["total"] or 0

    outstanding_balance = total_final_value - total_amount_paid

    total_platform_commission = requests_qs.aggregate(
        total=Sum("platform_commission_amount")
    )["total"] or 0

    total_technician_earnings = requests_qs.aggregate(
        total=Sum("technician_earning")
    )["total"] or 0

    total_paid_out = requests_qs.filter(
        technician_payout_status=ServiceRequest.PAYOUT_PAID
    ).aggregate(
        total=Sum("technician_earning")
    )["total"] or 0

    total_unpaid_payouts = total_technician_earnings - total_paid_out

    pending_payment_proofs = requests_qs.filter(
        payment_proof_status=ServiceRequest.PROOF_PENDING
    ).count()

    technician_report = (
        completed_requests
        .filter(assigned_technician__isnull=False)
        .values(
            "assigned_technician",
            "assigned_technician__user__username",
            "assigned_technician__business_name",
        )
        .annotate(
            completed_jobs=Count("id"),
            total_job_value=Sum("final_price"),
            total_earnings=Sum("technician_earning"),
            paid_out=Sum(
                "technician_earning",
                filter=models.Q(technician_payout_status=ServiceRequest.PAYOUT_PAID)
            ),
            unpaid=Sum(
                "technician_earning",
                filter=models.Q(technician_payout_status=ServiceRequest.PAYOUT_UNPAID)
            ),
        )
        .order_by("-total_earnings")
    )

    client_balances = (
        requests_qs
        .values(
            "client",
            "client__username",
            "client_name",
            "phone_number",
        )
        .annotate(
            total_requests=Count("id"),
            total_final_value=Sum("final_price"),
            total_paid=Sum("amount_paid"),
        )
        .order_by("-total_final_value")
    )

    client_balances_list = []

    for item in client_balances:
        final_value = item["total_final_value"] or 0
        paid = item["total_paid"] or 0
        item["balance"] = final_value - paid
        client_balances_list.append(item)

    return render(request, "admin_dashboard/reports.html", {
        "date_filter": date_filter,
        "start_date": start_date_input,
        "end_date": end_date_input,

        "total_requests": total_requests,
        "completed_jobs_count": completed_jobs_count,
        "cancelled_jobs_count": cancelled_jobs_count,

        "total_final_value": total_final_value,
        "total_amount_paid": total_amount_paid,
        "outstanding_balance": outstanding_balance,
        "total_platform_commission": total_platform_commission,
        "total_technician_earnings": total_technician_earnings,
        "total_paid_out": total_paid_out,
        "total_unpaid_payouts": total_unpaid_payouts,
        "pending_payment_proofs": pending_payment_proofs,

        "technician_report": technician_report,
        "client_balances": client_balances_list,
    })

@staff_member_required
def admin_reports_export_csv(request):
    today = timezone.localdate()

    date_filter = request.GET.get("date_filter", "this_month")
    start_date_input = request.GET.get("start_date", "")
    end_date_input = request.GET.get("end_date", "")

    start_date = None
    end_date = None

    if date_filter == "today":
        start_date = today
        end_date = today

    elif date_filter == "this_week":
        start_date = today - timedelta(days=today.weekday())
        end_date = today

    elif date_filter == "this_month":
        start_date = today.replace(day=1)
        end_date = today

    elif date_filter == "this_year":
        start_date = today.replace(month=1, day=1)
        end_date = today

    elif date_filter == "custom":
        if start_date_input:
            start_date = datetime.strptime(start_date_input, "%Y-%m-%d").date()

        if end_date_input:
            end_date = datetime.strptime(end_date_input, "%Y-%m-%d").date()

    requests_qs = ServiceRequest.objects.select_related(
        "category",
        "client",
        "assigned_technician",
        "assigned_technician__user",
    ).order_by("-created_at")

    if start_date:
        requests_qs = requests_qs.filter(created_at__date__gte=start_date)

    if end_date:
        requests_qs = requests_qs.filter(created_at__date__lte=end_date)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="fixnet_reports.csv"'

    writer = csv.writer(response)

    writer.writerow([
        "Request ID",
        "Title",
        "Client",
        "Phone",
        "Location",
        "Category",
        "Technician",
        "Status",
        "Payment Status",
        "Proof Status",
        "Final Price",
        "Amount Paid",
        "Balance Due",
        "Platform Commission",
        "Technician Earning",
        "Technician Payout Status",
        "Created At",
        "Completed At",
    ])

    for item in requests_qs:
        if item.client:
            client_name = item.client.username
        elif item.client_name:
            client_name = item.client_name
        else:
            client_name = "Public Client"

        technician_name = (
            item.assigned_technician.user.username
            if item.assigned_technician
            else "Not assigned"
        )

        category_name = item.category.name if item.category else "No category"

        writer.writerow([
            item.id,
            item.title,
            client_name,
            item.phone_number,
            item.location,
            category_name,
            technician_name,
            item.get_status_display(),
            item.get_payment_status_display(),
            item.get_payment_proof_status_display(),
            item.final_price or 0,
            item.amount_paid or 0,
            item.balance_due or 0,
            item.platform_commission_amount or 0,
            item.technician_earning or 0,
            item.get_technician_payout_status_display(),
            item.created_at.strftime("%Y-%m-%d %H:%M") if item.created_at else "",
            item.completed_at.strftime("%Y-%m-%d %H:%M") if item.completed_at else "",
        ])

    return response
@staff_member_required
def admin_reports_export_excel(request):
    today = timezone.localdate()

    date_filter = request.GET.get("date_filter", "this_month")
    start_date_input = request.GET.get("start_date", "")
    end_date_input = request.GET.get("end_date", "")

    start_date = None
    end_date = None

    if date_filter == "today":
        start_date = today
        end_date = today

    elif date_filter == "this_week":
        start_date = today - timedelta(days=today.weekday())
        end_date = today

    elif date_filter == "this_month":
        start_date = today.replace(day=1)
        end_date = today

    elif date_filter == "this_year":
        start_date = today.replace(month=1, day=1)
        end_date = today

    elif date_filter == "custom":
        if start_date_input:
            start_date = datetime.strptime(start_date_input, "%Y-%m-%d").date()

        if end_date_input:
            end_date = datetime.strptime(end_date_input, "%Y-%m-%d").date()

    requests_qs = ServiceRequest.objects.select_related(
        "category",
        "client",
        "assigned_technician",
        "assigned_technician__user",
    ).order_by("-created_at")

    if start_date:
        requests_qs = requests_qs.filter(created_at__date__gte=start_date)

    if end_date:
        requests_qs = requests_qs.filter(created_at__date__lte=end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "FixNet Reports"

    headers = [
        "Request ID",
        "Title",
        "Client",
        "Phone",
        "Location",
        "Category",
        "Technician",
        "Status",
        "Payment Status",
        "Proof Status",
        "Final Price",
        "Amount Paid",
        "Balance Due",
        "Platform Commission",
        "Technician Earning",
        "Technician Payout Status",
        "Created At",
        "Completed At",
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for item in requests_qs:
        if item.client:
            client_name = item.client.username
        elif item.client_name:
            client_name = item.client_name
        else:
            client_name = "Public Client"

        technician_name = (
            item.assigned_technician.user.username
            if item.assigned_technician
            else "Not assigned"
        )

        category_name = item.category.name if item.category else "No category"

        ws.append([
            item.id,
            item.title,
            client_name,
            item.phone_number,
            item.location,
            category_name,
            technician_name,
            item.get_status_display(),
            item.get_payment_status_display(),
            item.get_payment_proof_status_display(),
            float(item.final_price or 0),
            float(item.amount_paid or 0),
            float(item.balance_due or 0),
            float(item.platform_commission_amount or 0),
            float(item.technician_earning or 0),
            item.get_technician_payout_status_display(),
            item.created_at.strftime("%Y-%m-%d %H:%M") if item.created_at else "",
            item.completed_at.strftime("%Y-%m-%d %H:%M") if item.completed_at else "",
        ])

    # Make columns readable
    column_widths = {
        "A": 12,
        "B": 30,
        "C": 20,
        "D": 18,
        "E": 25,
        "F": 20,
        "G": 20,
        "H": 16,
        "I": 18,
        "J": 18,
        "K": 16,
        "L": 16,
        "M": 16,
        "N": 20,
        "O": 20,
        "P": 24,
        "Q": 20,
        "R": 20,
    }

    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="fixnet_reports.xlsx"'

    wb.save(response)
    return response

@staff_member_required
def admin_outstanding_balances_report(request):
    requests_qs = ServiceRequest.objects.select_related(
        "category",
        "client",
        "assigned_technician",
        "assigned_technician__user",
    ).order_by("-created_at")

    outstanding_requests = []

    total_outstanding = 0

    for item in requests_qs:
        balance = item.balance_due or 0

        if balance > 0:
            outstanding_requests.append({
                "request": item,
                "balance": balance,
            })
            total_outstanding += balance

    return render(request, "admin_dashboard/outstanding_balances.html", {
        "outstanding_requests": outstanding_requests,
        "total_outstanding": total_outstanding,
    })