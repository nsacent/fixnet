import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def build_excel_workbook(sheet_title, headers, rows, column_widths=None, metadata=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    current_row = 1

    if metadata:
        for label, value in metadata:
            ws.cell(row=current_row, column=1, value=label)
            ws.cell(row=current_row, column=2, value=value)

            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

        current_row += 1

    header_row = current_row
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in rows:
        ws.append(row)

    if column_widths:
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    return wb


def build_csv_response(filename, headers, rows, metadata=None):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    if metadata:
        for label, value in metadata:
            writer.writerow([label, value])

        writer.writerow([])

    writer.writerow(headers)

    for row in rows:
        writer.writerow(row)

    return response

def get_request_client_name(service_request):
    if service_request.client:
        return service_request.client.username

    if service_request.client_name:
        return service_request.client_name

    return "Public Client"


def get_request_technician_name(service_request):
    if service_request.assigned_technician:
        return service_request.assigned_technician.user.username

    return "Not assigned"


def get_request_category_name(service_request):
    if service_request.category:
        return service_request.category.name

    return "No category"


def format_datetime_for_export(value):
    if value:
        return value.strftime("%Y-%m-%d %H:%M")

    return ""


def build_request_export_row(service_request, include_completed_at=False, include_updated_at=False):
    row = [
        service_request.id,
        service_request.title,
        get_request_client_name(service_request),
        service_request.phone_number,
        service_request.location,
        get_request_category_name(service_request),
        get_request_technician_name(service_request),
        service_request.get_status_display(),
        service_request.get_payment_status_display(),
        service_request.get_payment_proof_status_display(),
        float(service_request.final_price or 0),
        float(service_request.amount_paid or 0),
        float(service_request.balance_due or 0),
        float(service_request.platform_commission_amount or 0),
        float(service_request.technician_earning or 0),
        service_request.get_technician_payout_status_display(),
        format_datetime_for_export(service_request.created_at),
    ]

    if include_completed_at:
        row.append(format_datetime_for_export(service_request.completed_at))

    if include_updated_at:
        row.append(format_datetime_for_export(service_request.updated_at))

    return row

def get_request_export_headers(include_completed_at=False, include_updated_at=False):
    headers = [
        "Request ID",
        "Title",
        "Client",
        "Phone",
        "Location",
        "Category",
        "Technician",
        "Status",
        "Payment Status",
        "Proof Status",
        "Final Price",
        "Amount Paid",
        "Balance Due",
        "Platform Commission",
        "Technician Earning",
        "Technician Payout Status",
        "Created At",
    ]

    if include_completed_at:
        headers.append("Completed At")

    if include_updated_at:
        headers.append("Updated At")

    return headers


def humanize_date_filter(value):
    labels = {
        "today": "Today",
        "this_week": "This Week",
        "this_month": "This Month",
        "this_year": "This Year",
        "custom": "Custom Range",
    }

    return labels.get(value, value or "This Month")


def build_report_metadata(date_data):
    return [
        ("Report Period", humanize_date_filter(date_data.get("date_filter", ""))),
        ("Start Date", date_data.get("start_date_input", "") or "Auto"),
        ("End Date", date_data.get("end_date_input", "") or "Auto"),
    ]